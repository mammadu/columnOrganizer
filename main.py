# -*- coding: utf-8 -*-
"""
Created on Thu Feb 13 22:13:18 2020

@author: muham
"""

from excelOrganizer import excelOrganizer
from openpyxl import load_workbook

# fileLocation = input("type in the file location")

file1 = 'Dover block Part numbers and lengths - New HP top cap.xlsx' #key file
file1Key ='Current carbon block assembly part number (w/current caps)' #this is the key column
file1Copy = 'New carbon block description' #this column contains the data to be copied

file2 = 'Phase 3 usage.xlsx' #file to be edited
file2Eval = 'Current Carbon block part number' #this is the column to be evaluated
file2Paste = 'New carbon block description' #this is the column that will be overwritten

#instantiating a class
eo = excelOrganizer()

#converting columns from excel files into dataframes.
c1 = eo.copyColumn(file1, file1Key) #dataframe of the key column
c2 = eo.copyColumn(file2, file2Eval) # dataframe of the column to be evaluated
n1 = eo.copyColumn(file1, file1Copy) #dataframe of the column to be copied
n2 = eo.copyColumn(file2, file2Paste) #dataframe column that will be overwritten

#if the value in (file1,carbonBlockColumn) equals the value in (file2,carbonBlockColumn) replace the value of (file2,newColumn) with the value of (file1,newColumn)
for i in range(0,len(c1)): #this loop is to go through every cell in in the first file
    for j in range(0,len(c2)): #this loop evaluates the value in the first file and then makes a change to the new column in the second file.
        if c1[i]==c2[j]:
            n2[j] = n1[i]


wb = load_workbook(file2) #this loads the workbook that will be edited
ws = wb['Phase III with usage'] #this specifies the worksheet in the workbook


for i in range(len(n2)):
    cell = 'G%d'  % (i + 2) #this specifies the specific cell to edit. The letter represents the column - %d is a placeholder for a number which is given by %()
    ws[cell] = n2[i]

wb.save(file2)

#to do
#   change the cell equation so that it already knows what column to copy over. Inputting the letter of the column in one location along with the column name is redundant
#   Make an intuituve GUI